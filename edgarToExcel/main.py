from bs4 import BeautifulSoup # actually the most beautiful things in the world
from openpyxl import Workbook
import req
import Parser


def isUniqueTitle(title,wb):
    """Check Title

    Params:
        title (String)  : The title to be looked for
        wb (Workbook)   : The workbook to look for title
    Return:
        Found (Boolean) : True if title is unique false otherwise
    """
    sheetnames = wb.sheetnames
    if(sheetnames.count(title) != 0):
        return False
    else:
        return True

def writeSheet(wb,title,table):
    if(isUniqueTitle(title,wb)):
        wb.create_sheet(title)
        wb.active = wb[title]
        ws = wb.active
        ws.append(["NAME OF ISSUER","TITLE OF CLASS","CUISP",'x$1000','PRN AMT','PRN','CALL','DISCRETION','SOLE'])
    else:
        wb.active = wb[title]
        ws = wb.active
    for entry in table:
        ws.append(entry)


def main():
    print("edgar To Excel V1")
    cik = "0001061768"
    allLink = req.getNumber(cik)
    wb = Workbook()

    for link in allLink:  #Get all of the text file
        textFile = req.getTextFile(link)
        soup = BeautifulSoup(textFile.content,features="html.parser")
        title = req.getDate(soup)
        if(Parser.isOldTextFormat(soup)):
            table = Parser.parseOldTextFileFormat(soup)
        else:
            table = Parser.parseNewTextFileFormat(soup)
        writeSheet(wb,title,table)

    wb.save('{companyCik}.xlsx'.format(companyCik = cik))

    print("end")

main()